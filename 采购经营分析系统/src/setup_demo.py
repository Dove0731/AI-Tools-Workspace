from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "01_原始数据"
CFG = ROOT / "02_配置文件"

ORDER_FIELDS = [
    "订单号", "订单日期", "物料编码", "物料名称", "物料类别", "规格型号", "计量单位",
    "采购数量", "含税单价", "税率", "未税单价", "含税金额", "供应商编码",
    "供应商名称", "采购员", "订单状态",
]
MASTER_FIELDS = ["物料编码", "标准物料名称", "物料类别", "标准规格", "标准计量单位", "品牌或生产商"]
QUOTE_FIELDS = ["报价日期", "物料编码", "供应商名称", "初始报价", "最终报价", "最小起订量", "付款条件", "报价有效期"]
RECEIPT_FIELDS = ["订单号", "物料编码", "入库日期", "实收数量", "入库单价", "批次号"]


def style_sheet(ws, widths=None):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if widths:
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
    else:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = 16
    if ws.max_row >= 2 and ws.max_column >= 1:
        tab = Table(displayName=f"T_{abs(hash(ws.title)) % 1000000}", ref=ws.dimensions)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(tab)


def save_book(path: Path, sheet_name: str, headers: list[str], rows: list[list]):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_templates():
    save_book(RAW / "本期采购订单" / "导入模板_采购订单.xlsx", "采购订单", ORDER_FIELDS, [])
    save_book(RAW / "历史采购订单" / "导入模板_历史采购订单.xlsx", "采购订单", ORDER_FIELDS, [])
    save_book(RAW / "供应商报价" / "导入模板_供应商报价.xlsx", "供应商报价", QUOTE_FIELDS, [])
    save_book(RAW / "入库记录" / "导入模板_入库记录.xlsx", "入库记录", RECEIPT_FIELDS, [])
    save_book(RAW / "物料主数据" / "导入模板_物料主数据.xlsx", "物料主数据", MASTER_FIELDS, [])


def build_mapping():
    aliases = {
        "订单号": "订单号;采购订单号;订单编号;订单ID;单据编号",
        "订单日期": "订单日期;采购日期;下单日期;制单日期",
        "物料编码": "物料编码;产品编码;商品编码;存货编码;物料编号",
        "物料名称": "物料名称;产品名称;商品名称;存货名称;原料名称",
        "物料类别": "物料类别;采购类别;品类;产品大类;材料类别",
        "规格型号": "规格型号;规格;型号;规格/尺寸;包材明细",
        "计量单位": "计量单位;单位;采购单位",
        "采购数量": "采购数量;下单数量;订单数量;数量",
        "含税单价": "含税单价;采购单价;采购成本;成交单价;单价",
        "税率": "税率;增值税率",
        "未税单价": "未税单价;不含税单价",
        "含税金额": "含税金额;采购金额;价税合计;订单金额",
        "供应商编码": "供应商编码;供应商编号",
        "供应商名称": "供应商名称;供应商;采购供应商",
        "采购员": "采购员;采购负责人;经办人",
        "订单状态": "订单状态;状态;单据状态",
    }
    wb = Workbook()
    wb.remove(wb.active)
    specs = [
        ("采购订单", ORDER_FIELDS),
        ("物料主数据", MASTER_FIELDS),
        ("供应商报价", QUOTE_FIELDS),
        ("入库记录", RECEIPT_FIELDS),
    ]
    extra_aliases = {
        "标准物料名称": "标准物料名称;物料名称;产品名称",
        "标准规格": "标准规格;规格型号;规格",
        "标准计量单位": "标准计量单位;计量单位;单位",
        "品牌或生产商": "品牌或生产商;品牌;生产商;制造商",
        "报价日期": "报价日期;询价日期;报价时间",
        "初始报价": "初始报价;首次报价;原报价",
        "最终报价": "最终报价;批准报价;谈判后报价",
        "最小起订量": "最小起订量;MOQ;起订量",
        "付款条件": "付款条件;账期;付款方式",
        "报价有效期": "报价有效期;有效期",
        "入库日期": "入库日期;收货日期;到货日期",
        "实收数量": "实收数量;入库数量;收货数量",
        "入库单价": "入库单价;收货单价",
        "批次号": "批次号;批号",
    }
    required = set(["订单号", "订单日期", "物料名称", "规格型号", "计量单位", "采购数量",
                    "供应商名称", "订单状态", "报价日期", "初始报价", "最终报价",
                    "入库日期", "实收数量", "标准物料名称", "标准规格", "标准计量单位"])
    for name, fields in specs:
        ws = wb.create_sheet(name)
        ws.append(["标准字段", "候选字段名（分号分隔）", "是否必需", "自动匹配说明"])
        for field in fields:
            cand = aliases.get(field, extra_aliases.get(field, field))
            note = "精确或别名匹配；多个候选同时出现时列入待确认"
            ws.append([field, cand, "是" if field in required else "否", note])
        style_sheet(ws, [18, 58, 12, 42])
    wb.save(CFG / "字段映射表.xlsx")


def build_units():
    rows = [
        ["千克", "克", 1000, "*", "标准数量=原数量×换算系数；标准单价=原单价÷换算系数"],
        ["kg", "克", 1000, "*", "通用质量换算"],
        ["公斤", "克", 1000, "*", "通用质量换算"],
        ["g", "克", 1, "*", "同单位"],
        ["克", "克", 1, "*", "同单位"],
        ["吨", "千克", 1000, "*", "通用质量换算"],
        ["t", "千克", 1000, "*", "通用质量换算"],
        ["件", "件", 1, "*", "同单位"],
        ["个", "个", 1, "*", "同单位"],
        ["套", "套", 1, "*", "同单位"],
        ["片", "片", 1, "*", "同单位"],
        ["支", "支", 1, "*", "同单位"],
        ["箱", "箱", 1, "*", "箱与个/件的换算必须按物料单独配置，当前不擅自换算"],
    ]
    save_book(CFG / "物料单位换算表.xlsx", "单位换算", ["原单位", "标准单位", "换算系数", "适用物料编码", "备注"], rows)


def build_rules():
    rows = [
        ["source_mode", "Excel", "数据源模式：Excel/ERP/Database"],
        ["price_basis", "含税", "价格对比统一口径：含税或未税"],
        ["exclude_status_keywords", "取消;已取消;作废;已作废;退货", "包含任一关键词的订单不参与分析"],
        ["last_price_increase_threshold", 0.05, "较上次采购价上涨阈值"],
        ["weighted_avg_increase_threshold", 0.08, "较12个月加权均价上涨阈值"],
        ["supplier_price_gap_threshold", 0.10, "同物料同期供应商价差阈值"],
        ["supplier_concentration_threshold", 0.80, "单一供应商采购金额占比阈值"],
        ["quantity_growth_threshold", 0.50, "采购数量环比异常增长阈值"],
        ["quote_overrun_tolerance", 0.0001, "实际价高于批准最终报价容差"],
        ["default_tax_rate", 0.13, "税率缺失时不自动填充，仅用于提示建议"],
        ["history_months", 12, "加权平均、最低价、最高价回溯月数"],
        ["candidate_name_similarity", 0.92, "无物料编码时，仅输出候选匹配的名称相似度阈值"],
        ["generate_pdf", "是", "支持时生成管理 PDF"],
        ["risk_high_amount", 10000, "异常影响金额达到该值时升级为高风险"],
        ["risk_medium_amount", 1000, "异常影响金额达到该值时升级为中风险"],
    ]
    save_book(CFG / "分析规则配置.xlsx", "规则配置", ["规则键", "规则值", "说明"], rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "名称标准化"
    ws.append(["对象类型", "原名称", "标准名称", "备注"])
    ws.append(["供应商", "广州清源原料有限公司", "广州清源原料有限公司", "示例；实际别名可继续追加"])
    ws.append(["物料", "透明质酸钠", "透明质酸钠", "示例"])
    style_sheet(ws, [16, 32, 32, 45])
    wb.save(CFG / "名称标准化表.xlsx")


def build_demo():
    master = [
        ["RM001", "透明质酸钠", "原料", "食品级/25kg", "克", "华东生物"],
        ["PK001", "30ml滴管瓶", "包材", "30ml/透明", "个", "粤美包材"],
        ["RM002", "烟酰胺", "原料", "化妆品级/25kg", "克", "维研"],
        ["FG001", "B5舒护面膜", "成品", "25ml×5片", "盒", "示范OEM"],
        ["OT001", "办公标签纸", "其他", "A4/50张", "包", "示范文具"],
    ]
    save_book(RAW / "物料主数据" / "演示数据_物料主数据.xlsx", "物料主数据", MASTER_FIELDS, master)

    history = [
        ["H250701", datetime(2025, 7, 15), "RM001", "透明质酸钠", "原料", "食品级/25kg", "千克", 100, 11.30, 0.13, 10.00, 1130, "S001", "广州清源原料有限公司", "李采购", "已完成"],
        ["H251001", datetime(2025, 10, 10), "RM001", "透明质酸钠", "原料", "食品级/25kg", "千克", 80, 11.00, 0.13, None, 880, "S001", "广州清源原料有限公司", "李采购", "已完成"],
        ["H260301", datetime(2026, 3, 8), "RM001", "透明质酸钠", "原料", "食品级/25kg", "千克", 120, 10.80, 0.13, None, 1296, "S002", "上海原素科技有限公司", "李采购", "已完成"],
        ["H251101", datetime(2025, 11, 6), "PK001", "30ml滴管瓶", "包材", "30ml/透明", "个", 50000, 0.80, 0.13, None, 40000, "S003", "粤美包材有限公司", "王采购", "已完成"],
        ["H260201", datetime(2026, 2, 18), "PK001", "30ml滴管瓶", "包材", "30ml/透明", "个", 60000, 0.82, 0.13, None, 49200, "S003", "粤美包材有限公司", "王采购", "已完成"],
        ["H260401", datetime(2026, 4, 12), "RM002", "烟酰胺", "原料", "化妆品级/25kg", "千克", 50, 33.00, 0.13, None, 1650, "S004", "维研原料有限公司", "李采购", "已完成"],
        ["H260501", datetime(2026, 5, 10), "RM002", "烟酰胺", "原料", "化妆品级/25kg", "千克", 50, 34.00, 0.13, None, 1700, "S004", "维研原料有限公司", "李采购", "已完成"],
        ["H260520", datetime(2026, 5, 20), "RM002", "烟酰胺", "原料", "化妆品级/25kg", "千克", 50, 35.00, 0.13, None, 1750, "S004", "维研原料有限公司", "李采购", "已完成"],
        ["H260599", datetime(2026, 5, 25), "PK001", "30ml滴管瓶", "包材", "30ml/透明", "个", 1000, 0.01, 0.13, None, 10, "S003", "粤美包材有限公司", "王采购", "已作废"],
    ]
    save_book(RAW / "历史采购订单" / "演示数据_历史采购订单.xlsx", "采购订单", ORDER_FIELDS, history)

    current = [
        ["C260601", datetime(2026, 6, 5), "RM001", "透明质酸钠", "原料", "食品级/25kg", "克", 120000, 0.01045, 0.13, None, 1254, "S001", "广州清源原料有限公司", "李采购", "已完成"],
        ["C260602", datetime(2026, 6, 8), "PK001", "30ml滴管瓶", "包材", "30ml/透明", "个", 100000, 0.90, 13, None, 90000, "S003", "粤美包材有限公司", "王采购", "已完成"],
        ["C260603", datetime(2026, 6, 9), "PK001", "30ml滴管瓶", "包材", "30ml/透明", "个", 30000, 0.81, 0.13, None, 24300, "S005", "华彩包装有限公司", "王采购", "已完成"],
        ["C260604", datetime(2026, 6, 15), "RM002", "烟酰胺", "原料", "化妆品级/25kg", "千克", 60, 36.00, 0.13, None, 2160, "S004", "维研原料有限公司", "李采购", "已完成"],
        ["C260605", datetime(2026, 6, 18), "FG001", "B5舒护面膜", "成品", "25ml×5片", "盒", 5000, 12.80, 0.13, None, 64000, "S006", "示范OEM工厂", "赵采购", "已完成"],
        ["C260606", datetime(2026, 6, 20), "", "透明质酸钠", "原料", "食品级/25kg", "千克", 10, 10.70, 0.13, None, 107, "S002", "上海原素科技有限公司", "李采购", "已完成"],
        ["C260607", datetime(2026, 6, 22), "OT001", "办公标签纸", "其他", "A4/50张", "箱", 2, 0, 0.13, 0, 0, "S007", "示范文具有限公司", "赵采购", "赠品"],
        ["C260608", datetime(2026, 6, 23), "PK001", "30ml滴管瓶", "包材", "30ml/透明", "个", 1000, 0.50, 0.13, None, 500, "S003", "粤美包材有限公司", "王采购", "已取消"],
    ]
    current.append(current[0].copy())  # 完全重复，测试去重与数据质量记录
    save_book(RAW / "本期采购订单" / "演示数据_本期采购订单.xlsx", "采购订单", ORDER_FIELDS, current)

    quotes = [
        [datetime(2026, 6, 1), "RM001", "广州清源原料有限公司", 0.01120, 0.01050, 100000, "月结30天", datetime(2026, 6, 30)],
        [datetime(2026, 6, 1), "PK001", "粤美包材有限公司", 0.88, 0.85, 50000, "月结30天", datetime(2026, 6, 30)],
        [datetime(2026, 6, 2), "PK001", "华彩包装有限公司", 0.86, 0.81, 30000, "预付30%", datetime(2026, 6, 30)],
        [datetime(2026, 6, 10), "RM002", "维研原料有限公司", 37.00, 35.50, 50, "月结45天", datetime(2026, 6, 30)],
        [datetime(2026, 6, 12), "FG001", "示范OEM工厂", 13.50, 12.80, 5000, "预付50%", datetime(2026, 7, 15)],
    ]
    save_book(RAW / "供应商报价" / "演示数据_供应商报价.xlsx", "供应商报价", QUOTE_FIELDS, quotes)

    receipts = [
        ["C260601", "RM001", datetime(2026, 6, 12), 120000, 0.01045, "DEMO-RM001-01"],
        ["C260602", "PK001", datetime(2026, 6, 18), 98000, 0.90, "DEMO-PK001-01"],
        ["C260603", "PK001", datetime(2026, 6, 19), 30000, 0.81, "DEMO-PK001-02"],
    ]
    save_book(RAW / "入库记录" / "演示数据_入库记录.xlsx", "入库记录", RECEIPT_FIELDS, receipts)


def main():
    for d in [RAW / "本期采购订单", RAW / "历史采购订单", RAW / "供应商报价",
              RAW / "入库记录", RAW / "物料主数据", CFG]:
        d.mkdir(parents=True, exist_ok=True)
    build_templates()
    build_mapping()
    build_units()
    build_rules()
    build_demo()
    print("模板、配置和明确标记的演示数据已创建。")


if __name__ == "__main__":
    main()

