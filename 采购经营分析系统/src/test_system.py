from __future__ import annotations

import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MONTH = "2026-06"
OUT = ROOT / "03_输出报告" / MONTH
LOG = ROOT / "05_运行日志"


def close(a, b, tol=0.01):
    return abs(float(a) - float(b)) <= tol


def latest_reports():
    return sorted(OUT.glob(f"采购月度分析报告_{MONTH}_V*.xlsx"))


def version(path: Path) -> int:
    return int(path.stem.rsplit("_V", 1)[1])


def run_once():
    result = subprocess.run(
        [sys.executable, str(ROOT / "src" / "main.py"), "--month", MONTH],
        cwd=ROOT, capture_output=True, text=True, encoding="utf-8", errors="replace"
    )
    if result.returncode != 0:
        raise AssertionError(f"月报运行失败：{result.stdout}\n{result.stderr}")


def main():
    before = latest_reports()
    run_once()
    after_first = latest_reports()
    run_once()
    after_second = latest_reports()
    assert len(after_first) == len(before) + 1, "首次测试运行未生成新版本"
    assert len(after_second) == len(after_first) + 1, "重复运行未生成新版本"
    assert version(after_second[-1]) == version(after_first[-1]) + 1, "版本号未递增"
    assert all(p.exists() for p in before), "重复运行覆盖或删除了旧报告"

    report = after_second[-1]
    compare = pd.read_excel(report, sheet_name="历史价格对比")
    savings = pd.read_excel(report, sheet_name="节约金额明细")
    anomalies = pd.read_excel(report, sheet_name="价格异常清单")
    current = pd.read_excel(report, sheet_name="本月采购明细")
    category = pd.read_excel(report, sheet_name="品类分析")
    supplier = pd.read_excel(report, sheet_name="供应商分析")

    rm1 = compare[compare["订单号"] == "C260601"].iloc[0]
    rm1_current = current[current["订单号"] == "C260601"].iloc[0]
    rm1_s = savings[savings["订单号"] == "C260601"].iloc[0]
    assert close(rm1_current["标准数量"], 120000), "千克/克单位换算失败"
    assert close(rm1["标准比较单价"], 0.01045, 0.000001), "标准单价换算失败"
    assert close(rm1["上次采购价"], 0.0108, 0.000001), "上次采购价匹配失败"
    assert close(rm1["近12个月加权平均价"], 0.01102, 0.000001), "12个月加权平均价失败"
    assert close(rm1_s["环比节约金额"], 42.00), "环比节约金额失败"
    assert close(rm1_s["历史基准节约金额"], 68.40), "历史基准节约金额失败"
    assert close(rm1_s["议价节约金额"], 90.00), "议价节约金额失败"

    pk1 = savings[savings["订单号"] == "C260602"].iloc[0]
    assert close(pk1["历史基准节约金额"], -8909.09, 0.02), "涨价基准计算失败"
    increase_sheet = pd.read_excel(report, sheet_name="涨价影响明细")
    pk1_inc = increase_sheet[increase_sheet["订单号"] == "C260602"].iloc[0]
    assert close(pk1_inc["涨价影响金额"], 8909.09, 0.02), "涨价影响金额失败"

    fg = compare[compare["订单号"] == "C260605"].iloc[0]
    assert fg["首次采购"] == "是", "首次采购识别失败"
    assert "连续三次涨价" in set(anomalies["异常类型"]), "连续三次涨价异常未识别"
    assert "同物料同期供应商价差" in set(anomalies["异常类型"]), "供应商同期价差异常未识别"
    assert "实际采购价高于批准报价" in set(anomalies["异常类型"]), "批准报价超支异常未识别"

    detail_total = current["含税金额"].fillna(0).sum()
    assert close(detail_total, category["采购金额"].fillna(0).sum()), "品类汇总与明细不一致"
    assert close(detail_total, supplier["采购金额"].fillna(0).sum()), "供应商汇总与明细不一致"

    wb = load_workbook(report, data_only=False)
    formula_count = sum(
        1 for ws in wb.worksheets for row in ws.iter_rows() for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    chart_count = sum(len(ws._charts) for ws in wb.worksheets)
    formula_errors = [
        f"{ws.title}!{cell.coordinate}" for ws in wb.worksheets for row in ws.iter_rows() for cell in row
        if isinstance(cell.value, str) and any(e in cell.value for e in ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"])
    ]
    assert formula_count >= len(savings) * 4, "节约复核公式数量不足"
    assert chart_count >= 3, "主报告图表数量不足"
    assert not formula_errors, f"发现公式错误：{formula_errors[:5]}"

    result = {
        "测试时间": datetime.now().isoformat(),
        "分析月份": MONTH,
        "测试版本": version(report),
        "测试结果": "全部通过",
        "验证项": [
            "历史价格匹配", "含税价与未税价转换", "计量单位换算", "首次采购识别",
            "环比节约金额", "12个月加权平均价", "涨价影响金额", "异常阈值判断",
            "重复运行版本保护", "Excel公式", "图表", "明细/品类/供应商汇总一致性",
        ],
        "关键结果": {
            "RM001_标准数量": rm1_current["标准数量"],
            "RM001_上次采购价": rm1["上次采购价"],
            "RM001_12个月加权均价": rm1["近12个月加权平均价"],
            "RM001_环比节约金额": rm1_s["环比节约金额"],
            "RM001_历史基准节约金额": rm1_s["历史基准节约金额"],
            "PK001_涨价影响金额": pk1_inc["涨价影响金额"],
            "采购总金额": detail_total,
            "公式数": formula_count,
            "图表数": chart_count,
            "异常数": len(anomalies),
        },
        "版本保护": {
            "测试前报告数": len(before),
            "第一次运行后": len(after_first),
            "第二次运行后": len(after_second),
            "最新版本": version(report),
        },
        "最新主报告": str(report),
    }
    path = LOG / f"端到端测试报告_{datetime.now():%Y%m%d_%H%M%S}.json"
    path.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
