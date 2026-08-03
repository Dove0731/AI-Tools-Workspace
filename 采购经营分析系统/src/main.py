from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import sys
import traceback
from collections import defaultdict
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "01_原始数据"
CFG = ROOT / "02_配置文件"
OUT = ROOT / "03_输出报告"
ARCHIVE = ROOT / "04_历史归档"
LOG_DIR = ROOT / "05_运行日志"

ORDER_FIELDS = [
    "订单号", "订单日期", "物料编码", "物料名称", "物料类别", "规格型号", "计量单位",
    "采购数量", "含税单价", "税率", "未税单价", "含税金额", "供应商编码",
    "供应商名称", "采购员", "订单状态",
]
MASTER_FIELDS = ["物料编码", "标准物料名称", "物料类别", "标准规格", "标准计量单位", "品牌或生产商"]
QUOTE_FIELDS = ["报价日期", "物料编码", "供应商名称", "初始报价", "最终报价", "最小起订量", "付款条件", "报价有效期"]
RECEIPT_FIELDS = ["订单号", "物料编码", "入库日期", "实收数量", "入库单价", "批次号"]

NAVY = "1F4E78"
BLUE = "D9EAF7"
PALE_BLUE = "EAF2F8"
GREEN = "E2F0D9"
GREEN_FONT = "006100"
RED = "FCE4D6"
RED_FONT = "9C0006"
AMBER = "FFF2CC"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
THIN_GRAY = Side(style="thin", color="D9E1F2")


class BusinessError(Exception):
    pass


def previous_month() -> str:
    today = pd.Timestamp.today().normalize()
    return (today.replace(day=1) - pd.Timedelta(days=1)).strftime("%Y-%m")


def valid_month(value: str) -> str:
    try:
        parsed = pd.Period(value, freq="M")
        return str(parsed)
    except Exception as exc:
        raise BusinessError(f"分析月份“{value}”格式错误，应为 YYYY-MM，例如 2026-06。") from exc


def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u3000", " ")).strip()


def norm_key(value) -> str:
    return re.sub(r"[\s\-_—/（）()]+", "", clean_text(value)).lower()


def to_number(value):
    if pd.isna(value) or clean_text(value) == "":
        return math.nan
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = clean_text(value).replace(",", "").replace("，", "").replace("￥", "").replace("¥", "")
    text = text.replace("%", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group()) if match else math.nan


def json_value(value):
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.isoformat()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return value


def load_rules() -> dict:
    path = CFG / "分析规则配置.xlsx"
    if not path.exists():
        raise BusinessError(f"缺少配置文件：{path}")
    df = pd.read_excel(path, sheet_name="规则配置")
    rules = {}
    for _, row in df.iterrows():
        key = clean_text(row.get("规则键"))
        if key:
            rules[key] = row.get("规则值")
    return rules


def numeric_rule(rules: dict, key: str, default: float) -> float:
    val = to_number(rules.get(key, default))
    return default if pd.isna(val) else float(val)


def load_mapping() -> dict:
    path = CFG / "字段映射表.xlsx"
    if not path.exists():
        raise BusinessError(f"缺少配置文件：{path}")
    xls = pd.ExcelFile(path)
    result = {}
    for sheet in xls.sheet_names:
        df = pd.read_excel(path, sheet_name=sheet)
        mapping = {}
        required = set()
        for _, row in df.iterrows():
            standard = clean_text(row.get("标准字段"))
            if not standard:
                continue
            aliases = [clean_text(x) for x in clean_text(row.get("候选字段名（分号分隔）")).split(";") if clean_text(x)]
            mapping[standard] = aliases + ([standard] if standard not in aliases else [])
            if clean_text(row.get("是否必需")) == "是":
                required.add(standard)
        result[sheet] = {"aliases": mapping, "required": required}
    return result


def load_unit_map() -> list[dict]:
    path = CFG / "物料单位换算表.xlsx"
    df = pd.read_excel(path, sheet_name="单位换算")
    rows = []
    for _, r in df.iterrows():
        factor = to_number(r.get("换算系数"))
        if pd.isna(factor) or factor <= 0:
            continue
        rows.append({
            "原单位": clean_text(r.get("原单位")),
            "标准单位": clean_text(r.get("标准单位")),
            "换算系数": float(factor),
            "适用物料编码": clean_text(r.get("适用物料编码")) or "*",
        })
    return rows


def load_name_map() -> dict:
    path = CFG / "名称标准化表.xlsx"
    if not path.exists():
        return {"供应商": {}, "物料": {}}
    df = pd.read_excel(path, sheet_name="名称标准化")
    result = {"供应商": {}, "物料": {}}
    for _, row in df.iterrows():
        kind = clean_text(row.get("对象类型"))
        raw = clean_text(row.get("原名称"))
        std = clean_text(row.get("标准名称"))
        if kind in result and raw and std:
            result[kind][norm_key(raw)] = std
    return result


def source_files(folder: Path) -> list[Path]:
    if not folder.exists():
        return []
    files = sorted(
        p for p in folder.glob("*.xlsx")
        if not p.name.startswith("~$") and "导入模板" not in p.name
    )
    real_files = [p for p in files if "演示数据" not in p.name]
    return real_files or files


def match_columns(columns, aliases: dict) -> tuple[dict, list[str]]:
    normalized = defaultdict(list)
    for col in columns:
        normalized[norm_key(col)].append(col)
    rename = {}
    problems = []
    for standard, names in aliases.items():
        hits = []
        for name in names:
            hits.extend(normalized.get(norm_key(name), []))
        hits = list(dict.fromkeys(hits))
        if len(hits) == 1:
            rename[hits[0]] = standard
        elif len(hits) > 1:
            problems.append(f"标准字段“{standard}”同时命中列：{', '.join(map(str, hits))}")
    return rename, problems


def read_dataset(folder: Path, dataset: str, mapping_cfg: dict, required_min: int = 2):
    rows = []
    issues = []
    manifest = []
    aliases = mapping_cfg[dataset]["aliases"]
    required = mapping_cfg[dataset]["required"]
    for path in source_files(folder):
        file_info = {"文件": str(path), "大小字节": path.stat().st_size, "工作表": []}
        try:
            xls = pd.ExcelFile(path)
        except Exception as exc:
            issues.append({"问题类型": "文件读取失败", "文件": str(path), "工作表": "", "数据行": "", "问题说明": str(exc), "建议动作": "检查文件是否损坏、加密或被占用", "人工确认状态": "待确认"})
            manifest.append(file_info)
            continue
        for sheet in xls.sheet_names:
            try:
                df = pd.read_excel(path, sheet_name=sheet, dtype=object)
            except Exception as exc:
                issues.append({"问题类型": "工作表读取失败", "文件": str(path), "工作表": sheet, "数据行": "", "问题说明": str(exc), "建议动作": "检查合并单元格、多行表头或文件格式", "人工确认状态": "待确认"})
                continue
            df = df.dropna(how="all")
            rename, ambiguous = match_columns(df.columns, aliases)
            score = len(rename)
            file_info["工作表"].append({"名称": sheet, "记录数": int(len(df)), "匹配字段数": score})
            if score < required_min:
                continue
            for msg in ambiguous:
                issues.append({"问题类型": "字段映射冲突", "文件": str(path), "工作表": sheet, "数据行": "", "问题说明": msg, "建议动作": "在字段映射表中删除冲突别名或调整原文件表头", "人工确认状态": "待确认"})
            missing = [x for x in required if x not in rename.values()]
            if missing:
                issues.append({"问题类型": "必需字段缺失", "文件": str(path), "工作表": sheet, "数据行": "", "问题说明": "缺少：" + "、".join(sorted(missing)), "建议动作": "补充字段或更新字段映射表；缺失数据不会被擅自推断", "人工确认状态": "待确认"})
            projected = df.rename(columns=rename)
            for field in aliases:
                if field not in projected.columns:
                    projected[field] = None
            projected = projected[list(aliases.keys())].copy()
            projected["_源文件"] = path.name
            projected["_源路径"] = str(path)
            projected["_源工作表"] = sheet
            projected["_源行号"] = projected.index + 2
            rows.append(projected)
        manifest.append(file_info)
    result = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame(columns=list(aliases.keys()) + ["_源文件", "_源路径", "_源工作表", "_源行号"])
    return result, issues, manifest


def find_unit_rule(unit: str, material_code: str, target_unit: str, unit_rules: list[dict]):
    if unit and target_unit and norm_key(unit) == norm_key(target_unit):
        return {"原单位": unit, "标准单位": target_unit, "换算系数": 1.0, "适用物料编码": "*"}
    exact = []
    generic = []
    for rule in unit_rules:
        if norm_key(rule["原单位"]) != norm_key(unit):
            continue
        if target_unit and norm_key(rule["标准单位"]) != norm_key(target_unit):
            continue
        scope = rule["适用物料编码"]
        if scope == material_code and material_code:
            exact.append(rule)
        elif scope == "*":
            generic.append(rule)
    candidates = exact or generic
    return candidates[0] if len(candidates) == 1 else None


def clean_master(master: pd.DataFrame, name_map: dict) -> pd.DataFrame:
    if master.empty:
        return master
    out = master.copy()
    for col in MASTER_FIELDS:
        out[col] = out[col].map(clean_text)
    out["标准物料名称"] = out["标准物料名称"].map(lambda x: name_map["物料"].get(norm_key(x), x))
    return out.drop_duplicates(subset=["物料编码"], keep="last")


def clean_orders(raw: pd.DataFrame, master: pd.DataFrame, unit_rules: list[dict], name_map: dict, rules: dict, source_kind: str):
    dq = []
    if raw.empty:
        return raw.copy(), dq, []
    df = raw.copy()
    df["_数据来源类型"] = source_kind
    for col in ["订单号", "物料编码", "物料名称", "物料类别", "规格型号", "计量单位", "供应商编码", "供应商名称", "采购员", "订单状态"]:
        df[col] = df[col].map(clean_text)
    df["供应商名称"] = df["供应商名称"].map(lambda x: name_map["供应商"].get(norm_key(x), x))
    df["物料名称"] = df["物料名称"].map(lambda x: name_map["物料"].get(norm_key(x), x))
    df["订单日期"] = pd.to_datetime(df["订单日期"], errors="coerce")
    for col in ["采购数量", "含税单价", "税率", "未税单价", "含税金额"]:
        df[col] = df[col].map(to_number)
    df["税率"] = df["税率"].map(lambda x: x / 100 if pd.notna(x) and x > 1 else x)
    df["_原始缺失字段"] = ""
    critical = ["订单号", "订单日期", "物料编码", "规格型号", "计量单位", "税率"]
    for idx, row in df.iterrows():
        missing = []
        for field in critical:
            val = row[field]
            if (field == "订单日期" and pd.isna(val)) or (field != "订单日期" and clean_text(val) == "") or (field == "税率" and pd.isna(val)):
                missing.append(field)
        df.at[idx, "_原始缺失字段"] = "、".join(missing)
        if missing:
            dq.append({"问题类型": "关键字段缺失", "文件": row["_源路径"], "工作表": row["_源工作表"], "数据行": row["_源行号"], "问题说明": "缺少：" + "、".join(missing), "建议动作": "补充原始字段；未补充前相关价格比较可能被阻断", "人工确认状态": "待确认"})
    duplicate_mask = df.duplicated(subset=ORDER_FIELDS, keep=False)
    duplicate_rows = df[duplicate_mask].copy()
    for _, row in duplicate_rows.iterrows():
        dq.append({"问题类型": "完全重复记录", "文件": row["_源路径"], "工作表": row["_源工作表"], "数据行": row["_源行号"], "问题说明": f"订单号 {row['订单号']} 的完全重复记录，清洗后只保留一条", "建议动作": "核对是否重复导出；系统已去重但不修改源文件", "人工确认状态": "待确认"})
    df = df.drop_duplicates(subset=ORDER_FIELDS, keep="first").reset_index(drop=True)

    exclude_words = [clean_text(x) for x in clean_text(rules.get("exclude_status_keywords", "取消;作废;退货")).split(";") if clean_text(x)]
    df["_排除原因"] = df["订单状态"].map(lambda s: next((w for w in exclude_words if w in s), ""))
    excluded = df[df["_排除原因"] != ""].copy()
    df = df[df["_排除原因"] == ""].copy()

    master_idx = master.set_index("物料编码").to_dict("index") if not master.empty else {}
    df["标准物料名称"] = ""
    df["标准规格"] = ""
    df["标准单位"] = ""
    df["品牌或生产商"] = ""
    df["候选物料编码"] = ""
    df["候选匹配相似度"] = math.nan
    threshold = numeric_rule(rules, "candidate_name_similarity", 0.92)
    for idx, row in df.iterrows():
        code = row["物料编码"]
        if code and code in master_idx:
            m = master_idx[code]
            df.at[idx, "标准物料名称"] = clean_text(m.get("标准物料名称")) or row["物料名称"]
            df.at[idx, "标准规格"] = clean_text(m.get("标准规格")) or row["规格型号"]
            df.at[idx, "标准单位"] = clean_text(m.get("标准计量单位")) or row["计量单位"]
            df.at[idx, "品牌或生产商"] = clean_text(m.get("品牌或生产商"))
            if not row["物料类别"]:
                df.at[idx, "物料类别"] = clean_text(m.get("物料类别"))
        else:
            df.at[idx, "标准物料名称"] = row["物料名称"]
            df.at[idx, "标准规格"] = row["规格型号"]
            if not code and not master.empty:
                target = norm_key(row["物料名称"] + row["规格型号"])
                scores = []
                for _, m in master.iterrows():
                    candidate = norm_key(clean_text(m["标准物料名称"]) + clean_text(m["标准规格"]))
                    score = SequenceMatcher(None, target, candidate).ratio() if target and candidate else 0
                    scores.append((score, clean_text(m["物料编码"]), clean_text(m["标准物料名称"])))
                scores.sort(reverse=True)
                if scores and scores[0][0] >= threshold:
                    df.at[idx, "候选物料编码"] = scores[0][1]
                    df.at[idx, "候选匹配相似度"] = scores[0][0]
                    dq.append({"问题类型": "无编码候选匹配", "文件": row["_源路径"], "工作表": row["_源工作表"], "数据行": row["_源行号"], "问题说明": f"候选 {scores[0][1]} {scores[0][2]}，相似度 {scores[0][0]:.1%}；系统未自动赋码", "建议动作": "由物料负责人确认编码后回填原始数据或映射配置", "人工确认状态": "待确认"})
    df["标准数量"] = math.nan
    df["标准含税单价"] = math.nan
    df["标准未税单价"] = math.nan
    df["单位换算系数"] = math.nan
    df["可比状态"] = "可比"
    for idx, row in df.iterrows():
        unit = row["计量单位"]
        target_unit = df.at[idx, "标准单位"]
        if not target_unit:
            target_unit = unit
            df.at[idx, "标准单位"] = target_unit
        rule = find_unit_rule(unit, row["物料编码"], target_unit, unit_rules)
        if not rule:
            df.at[idx, "可比状态"] = "不可比-缺少单位换算"
            dq.append({"问题类型": "缺少单位换算", "文件": row["_源路径"], "工作表": row["_源工作表"], "数据行": row["_源行号"], "问题说明": f"物料 {row['物料编码'] or '无编码'}：{unit or '空'} → {target_unit or '空'} 无唯一换算规则", "建议动作": "在物料单位换算表补充物料专属或通用换算后重跑", "人工确认状态": "待确认"})
            continue
        factor = rule["换算系数"]
        df.at[idx, "单位换算系数"] = factor
        qty = row["采购数量"]
        tax_price = row["含税单价"]
        ex_price = row["未税单价"]
        tax_rate = row["税率"]
        if pd.isna(tax_price) and pd.notna(ex_price) and pd.notna(tax_rate):
            tax_price = ex_price * (1 + tax_rate)
            df.at[idx, "含税单价"] = tax_price
        if pd.isna(ex_price) and pd.notna(tax_price) and pd.notna(tax_rate):
            ex_price = tax_price / (1 + tax_rate)
            df.at[idx, "未税单价"] = ex_price
        if pd.isna(df.at[idx, "含税金额"]) and pd.notna(tax_price) and pd.notna(qty):
            df.at[idx, "含税金额"] = tax_price * qty
        if pd.notna(qty):
            df.at[idx, "标准数量"] = qty * factor
        if pd.notna(tax_price):
            df.at[idx, "标准含税单价"] = tax_price / factor
        if pd.notna(ex_price):
            df.at[idx, "标准未税单价"] = ex_price / factor
        if pd.isna(qty) or (pd.isna(tax_price) and pd.isna(ex_price)):
            df.at[idx, "可比状态"] = "不可比-数量或价格缺失"
    basis = clean_text(rules.get("price_basis", "含税"))
    df["标准比较单价"] = df["标准含税单价"] if basis == "含税" else df["标准未税单价"]
    df["比较口径"] = basis
    df["_匹配键"] = df.apply(
        lambda r: "|".join([r["物料编码"], norm_key(r["标准规格"]), norm_key(r["标准单位"])])
        if r["物料编码"] and r["标准规格"] and r["标准单位"] and r["可比状态"] == "可比" else "", axis=1
    )
    df["赠品或零价格"] = df.apply(lambda r: "是" if (pd.notna(r["标准比较单价"]) and r["标准比较单价"] == 0) or "赠品" in r["订单状态"] else "否", axis=1)
    return df.reset_index(drop=True), dq, excluded.to_dict("records")


def clean_quotes(raw: pd.DataFrame, name_map: dict) -> pd.DataFrame:
    if raw.empty:
        return raw.copy()
    df = raw.copy()
    for col in ["物料编码", "供应商名称", "付款条件"]:
        df[col] = df[col].map(clean_text)
    df["供应商名称"] = df["供应商名称"].map(lambda x: name_map["供应商"].get(norm_key(x), x))
    for col in ["初始报价", "最终报价", "最小起订量"]:
        df[col] = df[col].map(to_number)
    for col in ["报价日期", "报价有效期"]:
        df[col] = pd.to_datetime(df[col], errors="coerce")
    return df


def calculate_baselines(current: pd.DataFrame, all_orders: pd.DataFrame, quotes: pd.DataFrame, rules: dict):
    months = int(numeric_rule(rules, "history_months", 12))
    rows = []
    all_valid = all_orders[
        (all_orders["可比状态"] == "可比") &
        (all_orders["_匹配键"] != "") &
        (all_orders["赠品或零价格"] == "否") &
        all_orders["订单日期"].notna() &
        all_orders["标准比较单价"].notna() &
        all_orders["标准数量"].notna()
    ].copy()
    for _, cur in current.sort_values(["订单日期", "订单号"]).iterrows():
        rec = cur.to_dict()
        rec.update({
            "上次采购价": math.nan, "近12个月加权平均价": math.nan, "近12个月最低价": math.nan,
            "近12个月最高价": math.nan, "历史有效记录数": 0, "环比节约金额": math.nan,
            "历史基准节约金额": math.nan, "涨价影响金额": 0.0, "初始报价": math.nan,
            "最终批准报价": math.nan, "议价节约金额": math.nan, "节约率": math.nan,
            "首次采购": "否", "规格变化": "否", "单位变化": "否", "人工确认状态": "待确认",
            "正式汇总节约金额": 0.0,
        })
        key = cur["_匹配键"]
        date = cur["订单日期"]
        if key and pd.notna(date) and cur["可比状态"] == "可比" and cur["赠品或零价格"] == "否":
            prior = all_valid[(all_valid["_匹配键"] == key) & (all_valid["订单日期"] < date)].sort_values("订单日期")
            window_start = date - pd.DateOffset(months=months)
            window = prior[prior["订单日期"] >= window_start]
            if not prior.empty:
                rec["上次采购价"] = float(prior.iloc[-1]["标准比较单价"])
            if not window.empty and window["标准数量"].sum() > 0:
                total_qty = float(window["标准数量"].sum())
                weighted = float((window["标准比较单价"] * window["标准数量"]).sum() / total_qty)
                rec["近12个月加权平均价"] = weighted
                rec["近12个月最低价"] = float(window["标准比较单价"].min())
                rec["近12个月最高价"] = float(window["标准比较单价"].max())
                rec["历史有效记录数"] = int(len(window))
            if pd.isna(rec["上次采购价"]) and pd.isna(rec["近12个月加权平均价"]):
                rec["首次采购"] = "是"
            qty = cur["标准数量"]
            price = cur["标准比较单价"]
            if pd.notna(rec["上次采购价"]):
                rec["环比节约金额"] = (rec["上次采购价"] - price) * qty
            if pd.notna(rec["近12个月加权平均价"]):
                hist_saving = (rec["近12个月加权平均价"] - price) * qty
                rec["历史基准节约金额"] = hist_saving
                rec["涨价影响金额"] = max(-hist_saving, 0)
                base_amount = rec["近12个月加权平均价"] * qty
                rec["节约率"] = hist_saving / base_amount if base_amount else math.nan
        elif cur["赠品或零价格"] == "是":
            rec["首次采购"] = "不适用"

        same_code_prior = all_orders[
            (all_orders["物料编码"] == cur["物料编码"]) &
            (all_orders["订单日期"] < date)
        ] if cur["物料编码"] and pd.notna(date) else pd.DataFrame()
        if not same_code_prior.empty:
            if norm_key(cur["标准规格"]) not in set(same_code_prior["标准规格"].map(norm_key)):
                rec["规格变化"] = "是"
            if norm_key(cur["标准单位"]) not in set(same_code_prior["标准单位"].map(norm_key)):
                rec["单位变化"] = "是"

        if not quotes.empty and cur["物料编码"] and pd.notna(date):
            candidates = quotes[
                (quotes["物料编码"] == cur["物料编码"]) &
                (quotes["供应商名称"] == cur["供应商名称"]) &
                (quotes["报价日期"] <= date) &
                (quotes["报价有效期"].isna() | (quotes["报价有效期"] >= date))
            ].sort_values("报价日期")
            if not candidates.empty:
                q = candidates.iloc[-1]
                rec["初始报价"] = q["初始报价"]
                rec["最终批准报价"] = q["最终报价"]
                factor = cur["单位换算系数"] if pd.notna(cur["单位换算系数"]) else 1
                if pd.notna(q["初始报价"]) and pd.notna(cur["标准数量"]):
                    rec["议价节约金额"] = (q["初始报价"] / factor - cur["标准比较单价"]) * cur["标准数量"]
        rows.append(rec)
    return pd.DataFrame(rows)


def risk_level(amount: float, rules: dict, inherent: str = "中") -> str:
    amount = abs(amount) if pd.notna(amount) else 0
    if amount >= numeric_rule(rules, "risk_high_amount", 10000):
        return "高"
    if amount >= numeric_rule(rules, "risk_medium_amount", 1000):
        return "中"
    return inherent


def anomaly_row(row, kind: str, basis: str, amount: float, action: str, rules: dict, inherent="中"):
    return {
        "订单号": row.get("订单号", ""),
        "订单日期": row.get("订单日期", pd.NaT),
        "物料编码": row.get("物料编码", ""),
        "物料名称": row.get("标准物料名称", row.get("物料名称", "")),
        "供应商名称": row.get("供应商名称", ""),
        "异常类型": kind,
        "判断依据": basis,
        "影响金额": float(amount) if pd.notna(amount) else 0.0,
        "风险等级": risk_level(amount, rules, inherent),
        "建议处理动作": action,
        "人工确认状态": "待确认",
    }


def detect_anomalies(detail: pd.DataFrame, all_orders: pd.DataFrame, duplicate_dq: list[dict], rules: dict):
    anomalies = []
    last_th = numeric_rule(rules, "last_price_increase_threshold", 0.05)
    avg_th = numeric_rule(rules, "weighted_avg_increase_threshold", 0.08)
    quote_tol = numeric_rule(rules, "quote_overrun_tolerance", 0.0001)
    for _, row in detail.iterrows():
        price = row["标准比较单价"]
        qty = row["标准数量"]
        last = row["上次采购价"]
        avg = row["近12个月加权平均价"]
        if pd.notna(last) and last > 0 and pd.notna(price) and price / last - 1 > last_th:
            delta = (price - last) * qty
            anomalies.append(anomaly_row(row, "较上次采购价上涨", f"本次 {price:.6f}，上次 {last:.6f}，涨幅 {price / last - 1:.1%} > {last_th:.1%}", delta, "复核涨价原因、替代供应商和议价记录；必要时暂停未执行数量", rules))
        if pd.notna(avg) and avg > 0 and pd.notna(price) and price / avg - 1 > avg_th:
            delta = (price - avg) * qty
            anomalies.append(anomaly_row(row, "高于12个月加权均价", f"本次 {price:.6f}，12个月均价 {avg:.6f}，高出 {price / avg - 1:.1%} > {avg_th:.1%}", delta, "核对规格、税价和市场变动，补充审批依据", rules))
        approved = row["最终批准报价"]
        factor = row["单位换算系数"] if pd.notna(row["单位换算系数"]) else 1
        if pd.notna(approved) and pd.notna(price) and price > approved / factor + quote_tol:
            delta = (price - approved / factor) * qty
            anomalies.append(anomaly_row(row, "实际采购价高于批准报价", f"本次 {price:.6f}，批准报价折算 {approved / factor:.6f}", delta, "核对成交审批、订单单价和报价有效期；追回未经批准的价差", rules, "高"))
        if clean_text(row["_原始缺失字段"]):
            anomalies.append(anomaly_row(row, "关键字段缺失", row["_原始缺失字段"], 0, "补齐物料编码、规格、单位或税率后重跑", rules, "中"))
        if row["首次采购"] == "是":
            anomalies.append(anomaly_row(row, "首次采购或无有效历史价", "相同物料编码+规格+统一单位无有效历史记录", 0, "至少完成三家比价或书面说明单一来源，并由经营负责人确认", rules, "中"))
        if row["规格变化"] == "是" or row["单位变化"] == "是":
            anomalies.append(anomaly_row(row, "规格或单位变化", f"规格变化={row['规格变化']}；单位变化={row['单位变化']}", 0, "确认是否为同一可比物料；缺少转换规则时不得计算节约", rules, "中"))
        if row["_匹配键"] and pd.notna(row["订单日期"]) and row["可比状态"] == "可比":
            series = all_orders[
                (all_orders["_匹配键"] == row["_匹配键"]) &
                (all_orders["订单日期"] <= row["订单日期"]) &
                all_orders["标准比较单价"].notna() &
                (all_orders["赠品或零价格"] == "否")
            ].sort_values(["订单日期", "订单号"]).tail(4)["标准比较单价"].tolist()
            if len(series) >= 4 and all(series[i] > series[i - 1] for i in range(1, 4)):
                impact = (series[-1] - series[-2]) * row["标准数量"]
                anomalies.append(anomaly_row(row, "连续三次涨价", "最近4次价格依次为：" + " → ".join(f"{x:.6f}" for x in series), impact, "升级采购负责人复核，要求供应商提交成本变动证据并启动替代询价", rules, "高"))

    gap_th = numeric_rule(rules, "supplier_price_gap_threshold", 0.10)
    comparable = detail[(detail["_匹配键"] != "") & detail["标准比较单价"].notna()]
    for key, group in comparable.groupby("_匹配键"):
        if group["供应商名称"].nunique() < 2:
            continue
        min_price = group["标准比较单价"].min()
        max_price = group["标准比较单价"].max()
        if min_price > 0 and max_price / min_price - 1 > gap_th:
            high_rows = group[group["标准比较单价"] == max_price]
            for _, row in high_rows.iterrows():
                amount = (max_price - min_price) * row["标准数量"]
                anomalies.append(anomaly_row(row, "同物料同期供应商价差", f"最高 {max_price:.6f}，最低 {min_price:.6f}，价差 {max_price / min_price - 1:.1%} > {gap_th:.1%}", amount, "复核质量、交期、账期差异；条件可比时优先向低价供应商分配增量", rules))

    conc_th = numeric_rule(rules, "supplier_concentration_threshold", 0.80)
    for code, group in detail[detail["物料编码"] != ""].groupby("物料编码"):
        total = group["含税金额"].fillna(0).sum()
        if total <= 0:
            continue
        by_supplier = group.groupby("供应商名称")["含税金额"].sum().sort_values(ascending=False)
        supplier = by_supplier.index[0]
        share = by_supplier.iloc[0] / total
        if share > conc_th:
            row = group[group["供应商名称"] == supplier].iloc[0]
            anomalies.append(anomaly_row(row, "单一供应商集中度过高", f"{supplier} 占物料本月采购额 {share:.1%} > {conc_th:.1%}", by_supplier.iloc[0], "评估第二供应源、切换成本和安全库存；非价格因素需形成书面说明", rules, "中"))

    qty_th = numeric_rule(rules, "quantity_growth_threshold", 0.50)
    for key, group in comparable.groupby("_匹配键"):
        row = group.iloc[-1]
        month_start = row["订单日期"].to_period("M").start_time
        prev_start = month_start - pd.DateOffset(months=1)
        hist = all_orders[(all_orders["_匹配键"] == key) & (all_orders["订单日期"] >= prev_start) & (all_orders["订单日期"] < month_start)]
        prev_qty = hist["标准数量"].sum()
        cur_qty = group["标准数量"].sum()
        if prev_qty > 0 and cur_qty / prev_qty - 1 > qty_th:
            prev_avg = (hist["标准比较单价"] * hist["标准数量"]).sum() / prev_qty
            cur_avg = (group["标准比较单价"] * group["标准数量"]).sum() / cur_qty
            if cur_avg >= prev_avg:
                anomalies.append(anomaly_row(row, "采购数量异常增长但单价未下降", f"数量环比增长 {cur_qty / prev_qty - 1:.1%}，单价 {cur_avg:.6f} 未低于上月 {prev_avg:.6f}", (cur_avg - prev_avg) * cur_qty, "核对需求与库存，重新谈判阶梯价格和MOQ", rules))

    for item in duplicate_dq:
        if item["问题类型"] == "完全重复记录":
            dummy = {"订单号": re.search(r"订单号 (.*?) 的", item["问题说明"]).group(1) if "订单号 " in item["问题说明"] else "", "订单日期": pd.NaT, "物料编码": "", "标准物料名称": "", "供应商名称": ""}
            anomalies.append(anomaly_row(dummy, "疑似重复订单", item["问题说明"], 0, "核对导出范围和订单唯一键；系统已删除完全重复行", rules, "中"))
    return pd.DataFrame(anomalies)


def category_and_supplier(detail: pd.DataFrame):
    category = detail.groupby("物料类别", dropna=False).agg(
        采购金额=("含税金额", "sum"), 订单数=("订单号", "nunique"),
        物料数=("物料编码", lambda x: x[x != ""].nunique()), 采购数量=("标准数量", "sum")
    ).reset_index()
    category["金额占比"] = category["采购金额"] / category["采购金额"].sum() if category["采购金额"].sum() else 0
    supplier = detail.groupby("供应商名称", dropna=False).agg(
        采购金额=("含税金额", "sum"), 订单数=("订单号", "nunique"),
        物料数=("物料编码", lambda x: x[x != ""].nunique()),
        历史基准节约金额=("历史基准节约金额", "sum"), 涨价影响金额=("涨价影响金额", "sum")
    ).reset_index().sort_values("采购金额", ascending=False)
    return category, supplier


def monthly_trend(all_orders: pd.DataFrame, month: str):
    end = pd.Period(month, "M").end_time
    start = (pd.Period(month, "M") - 11).start_time
    df = all_orders[
        (all_orders["订单日期"] >= start) & (all_orders["订单日期"] <= end) &
        all_orders["含税金额"].notna()
    ].copy()
    df["月份"] = df["订单日期"].dt.to_period("M").astype(str)
    grouped = df.groupby("月份").agg(采购金额=("含税金额", "sum"), 标准数量=("标准数量", "sum")).reset_index()
    avg_rows = []
    for period, g in df.groupby("月份"):
        valid = g[g["标准比较单价"].notna() & g["标准数量"].notna()]
        qty = valid["标准数量"].sum()
        avg_price = (valid["标准比较单价"] * valid["标准数量"]).sum() / qty if qty else math.nan
        avg_rows.append((period, avg_price))
    avg_map = dict(avg_rows)
    grouped["跨物料加权平均单价_仅辅助"] = grouped["月份"].map(avg_map)
    periods = [str(pd.Period(month, "M") - i) for i in range(11, -1, -1)]
    return pd.DataFrame({"月份": periods}).merge(grouped, on="月份", how="left").fillna({"采购金额": 0, "标准数量": 0})


def display_df(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    out = df.copy()
    for col in columns:
        if col not in out.columns:
            out[col] = None
    return out[columns]


def safe_excel_value(value):
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, (pd.Int64Dtype,)):
        return int(value)
    return value


def write_table_sheet(wb: Workbook, title: str, df: pd.DataFrame, table_name: str, freeze="A2"):
    ws = wb.create_sheet(title)
    if df.empty and len(df.columns) == 0:
        df = pd.DataFrame({"说明": ["本期无记录"]})
    headers = list(df.columns)
    ws.append(headers)
    for row in df.itertuples(index=False, name=None):
        ws.append([safe_excel_value(v) for v in row])
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    for c, header in enumerate(headers, 1):
        values = [clean_text(ws.cell(r, c).value) for r in range(1, min(ws.max_row, 200) + 1)]
        width = min(max(max((len(v) for v in values), default=8) + 2, 11), 34)
        if any(k in header for k in ["说明", "依据", "动作", "路径"]):
            width = min(max(width, 26), 48)
        ws.column_dimensions[get_column_letter(c)].width = width
        if any(k in header for k in ["日期"]):
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = "yyyy-mm-dd"
        if any(k in header for k in ["金额", "单价", "报价", "采购价", "均价", "最低价", "最高价"]):
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = '#,##0.00;[Red](#,##0.00);-'
        if any(k in header for k in ["率", "占比", "相似度"]):
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = "0.0%"
    if ws.max_row >= 2 and ws.max_column >= 1:
        table = Table(displayName=table_name, ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)
    return ws


def add_conditional_formats(ws):
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for name in ["环比节约金额", "历史基准节约金额", "议价节约金额", "正式汇总节约金额"]:
        if name in headers and ws.max_row >= 2:
            col = get_column_letter(headers[name])
            ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_FONT)))
    for name in ["涨价影响金额", "影响金额"]:
        if name in headers and ws.max_row >= 2:
            col = get_column_letter(headers[name])
            ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_FONT)))
    if "风险等级" in headers and ws.max_row >= 2:
        col = get_column_letter(headers["风险等级"])
        ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", FormulaRule(formula=[f'${col}2="高"'], fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_FONT, bold=True)))


def build_dashboard(wb: Workbook, month: str, version: int, detail: pd.DataFrame, anomalies: pd.DataFrame, category: pd.DataFrame, trend: pd.DataFrame):
    ws = wb.create_sheet("管理驾驶舱", 0)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H2")
    ws["A1"] = f"采购月度管理驾驶舱｜{month}"
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"] = f"版本 V{version:03d}｜生成时间 {datetime.now():%Y-%m-%d %H:%M}｜正式节约口径：近12个月加权平均价，经人工确认后计入"
    ws.merge_cells("A3:H3")
    ws["A3"].font = Font(color="666666", italic=True)
    kpis = [
        ("本月订单数", int(detail["订单号"].nunique())),
        ("采购总金额", float(detail["含税金额"].fillna(0).sum())),
        ("采购物料数", int(detail.loc[detail["物料编码"] != "", "物料编码"].nunique())),
        ("活跃供应商数", int(detail["供应商名称"].replace("", pd.NA).nunique())),
        ("经确认节约金额", 0.0),
        ("节约率", 0.0),
        ("涨价影响金额", float(detail["涨价影响金额"].fillna(0).sum())),
        ("异常记录数", int(len(anomalies))),
    ]
    positions = ["A5", "C5", "E5", "G5", "A8", "C8", "E8", "G8"]
    for (label, value), pos in zip(kpis, positions):
        cell = ws[pos]
        col = cell.column
        row = cell.row
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 1)
        ws.cell(row, col).value = label
        ws.cell(row, col).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(row, col).font = Font(bold=True, color=NAVY)
        ws.cell(row, col).alignment = Alignment(horizontal="center")
        val_cell = ws.cell(row + 1, col)
        val_cell.value = value
        val_cell.font = Font(size=16, bold=True, color=RED_FONT if label in ["涨价影响金额", "异常记录数"] else GREEN_FONT if label == "经确认节约金额" else NAVY)
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        if "金额" in label:
            val_cell.number_format = '#,##0.00;[Red](#,##0.00);-'
        if "率" in label:
            val_cell.number_format = "0.0%"
        for rr in range(row, row + 3):
            for cc in range(col, col + 2):
                ws.cell(rr, cc).border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
    savings_end_row = len(detail) + 1
    ws["A9"] = f"=SUM('节约金额明细'!R2:R{savings_end_row})"
    ws["C9"] = f"=IFERROR(A9/SUMPRODUCT('节约金额明细'!J2:J{savings_end_row},'节约金额明细'!G2:G{savings_end_row}),0)"
    ws["A14"] = "节约贡献 TOP10 物料（未确认，仅分析值）"
    ws["E14"] = "涨价影响 TOP10 物料"
    for c in ["A14", "E14"]:
        ws[c].font = Font(bold=True, color=WHITE)
        ws[c].fill = PatternFill("solid", fgColor=NAVY)
    savings_top = detail.groupby(["物料编码", "标准物料名称"], dropna=False)["历史基准节约金额"].sum().reset_index()
    savings_top = savings_top[savings_top["历史基准节约金额"] > 0].sort_values("历史基准节约金额", ascending=False).head(10)
    increase_top = detail.groupby(["物料编码", "标准物料名称"], dropna=False)["涨价影响金额"].sum().reset_index()
    increase_top = increase_top[increase_top["涨价影响金额"] > 0].sort_values("涨价影响金额", ascending=False).head(10)
    ws.append([])
    for idx, headers in [(1, ["物料", "节约金额"]), (5, ["物料", "涨价影响"])]:
        ws.cell(15, idx).value, ws.cell(15, idx + 1).value = headers
        for cc in [idx, idx + 1]:
            ws.cell(15, cc).fill = PatternFill("solid", fgColor=GRAY)
            ws.cell(15, cc).font = Font(bold=True)
    for i, (_, r) in enumerate(savings_top.iterrows(), 16):
        ws.cell(i, 1).value = f"{r['物料编码']} {r['标准物料名称']}"
        ws.cell(i, 2).value = r["历史基准节约金额"]
        ws.cell(i, 2).number_format = '#,##0.00'
    for i, (_, r) in enumerate(increase_top.iterrows(), 16):
        ws.cell(i, 5).value = f"{r['物料编码']} {r['标准物料名称']}"
        ws.cell(i, 6).value = r["涨价影响金额"]
        ws.cell(i, 6).number_format = '#,##0.00'

    cat_start = 1
    ws.cell(cat_start, 10).value = "品类"
    ws.cell(cat_start, 11).value = "采购金额"
    for i, (_, r) in enumerate(category.sort_values("采购金额", ascending=False).iterrows(), cat_start + 1):
        ws.cell(i, 10).value = r["物料类别"] or "未分类"
        ws.cell(i, 11).value = r["采购金额"]
    trend_start = 1
    ws.cell(trend_start, 13).value = "月份"
    ws.cell(trend_start, 14).value = "采购金额"
    ws.cell(trend_start, 15).value = "跨物料加权平均单价"
    for i, (_, r) in enumerate(trend.iterrows(), trend_start + 1):
        ws.cell(i, 13).value = r["月份"]
        ws.cell(i, 14).value = r["采购金额"]
        ws.cell(i, 15).value = None if pd.isna(r["跨物料加权平均单价_仅辅助"]) else r["跨物料加权平均单价_仅辅助"]

    if not category.empty:
        pie = PieChart()
        pie.title = "各品类采购金额占比"
        pie.add_data(Reference(ws, min_col=11, min_row=cat_start, max_row=cat_start + len(category)), titles_from_data=True)
        pie.set_categories(Reference(ws, min_col=10, min_row=cat_start + 1, max_row=cat_start + len(category)))
        pie.height, pie.width = 7, 11
        ws.add_chart(pie, "A28")
    line = LineChart()
    line.title = "近12个月采购金额趋势（含税）"
    line.y_axis.title = "金额"
    line.x_axis.title = "月份"
    line.add_data(Reference(ws, min_col=14, min_row=trend_start, max_row=trend_start + len(trend)), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=13, min_row=trend_start + 1, max_row=trend_start + len(trend)))
    line.height, line.width = 7, 12
    ws.add_chart(line, "E28")
    price_line = LineChart()
    price_line.title = "近12个月跨物料加权平均单价（仅辅助观察）"
    price_line.y_axis.title = "统一税价口径单价"
    price_line.x_axis.title = "月份"
    price_line.add_data(Reference(ws, min_col=15, min_row=trend_start, max_row=trend_start + len(trend)), titles_from_data=True)
    price_line.set_categories(Reference(ws, min_col=13, min_row=trend_start + 1, max_row=trend_start + len(trend)))
    price_line.height, price_line.width = 7, 12
    ws.add_chart(price_line, "A44")
    for col, width in {"A": 24, "B": 16, "C": 18, "D": 4, "E": 24, "F": 16, "G": 18, "H": 4}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"
    return ws


def add_savings_formulas(ws):
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    needed = ["标准数量", "标准比较单价", "上次采购价", "近12个月加权平均价", "人工确认状态"]
    if not all(x in headers for x in needed):
        return
    new_headers = ["复核_环比节约金额", "复核_历史基准节约金额", "复核_涨价影响金额"]
    start = ws.max_column + 1
    for j, h in enumerate(new_headers, start):
        ws.cell(1, j).value = h
        ws.cell(1, j).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(1, j).font = Font(color=WHITE, bold=True)
        ws.column_dimensions[get_column_letter(j)].width = 22
    q = get_column_letter(headers["标准数量"])
    p = get_column_letter(headers["标准比较单价"])
    lp = get_column_letter(headers["上次采购价"])
    wa = get_column_letter(headers["近12个月加权平均价"])
    cf = get_column_letter(headers["人工确认状态"])
    for r in range(2, ws.max_row + 1):
        ws.cell(r, start).value = f'=IF(OR({lp}{r}="",{p}{r}="",{q}{r}=""),"",({lp}{r}-{p}{r})*{q}{r})'
        ws.cell(r, start + 1).value = f'=IF(OR({wa}{r}="",{p}{r}="",{q}{r}=""),"",({wa}{r}-{p}{r})*{q}{r})'
        ws.cell(r, start + 2).value = f'=IF(OR({wa}{r}="",{p}{r}="",{q}{r}=""),"",MAX(({p}{r}-{wa}{r})*{q}{r},0))'
        official = headers.get("正式汇总节约金额")
        if official:
            ws.cell(r, official).value = f'=IF({cf}{r}="已确认",MAX({get_column_letter(start + 1)}{r},0),0)'
        for c in range(start, start + 3):
            ws.cell(r, c).number_format = '#,##0.00;[Red](#,##0.00);-'
    ws.auto_filter.ref = ws.dimensions


def build_main_report(path: Path, month: str, version: int, detail: pd.DataFrame, all_orders: pd.DataFrame, anomalies: pd.DataFrame, dq: pd.DataFrame, category: pd.DataFrame, supplier: pd.DataFrame, trend: pd.DataFrame, rules: dict):
    wb = Workbook()
    wb.remove(wb.active)
    build_dashboard(wb, month, version, detail, anomalies, category, trend)
    current_cols = ["订单号", "订单日期", "物料编码", "标准物料名称", "物料类别", "标准规格", "计量单位", "标准单位", "采购数量", "标准数量", "含税单价", "未税单价", "税率", "含税金额", "供应商名称", "采购员", "订单状态", "赠品或零价格", "可比状态", "_源文件", "_源工作表", "_源行号"]
    hist_cols = ["订单号", "订单日期", "物料编码", "标准物料名称", "标准规格", "标准单位", "供应商名称", "标准比较单价", "上次采购价", "近12个月加权平均价", "近12个月最低价", "近12个月最高价", "历史有效记录数", "首次采购", "规格变化", "单位变化", "可比状态"]
    saving_cols = ["订单号", "订单日期", "物料编码", "标准物料名称", "供应商名称", "标准单位", "标准数量", "标准比较单价", "上次采购价", "近12个月加权平均价", "初始报价", "最终批准报价", "环比节约金额", "历史基准节约金额", "议价节约金额", "节约率", "人工确认状态", "正式汇总节约金额"]
    increase_cols = ["订单号", "订单日期", "物料编码", "标准物料名称", "供应商名称", "标准数量", "标准比较单价", "近12个月加权平均价", "涨价影响金额", "人工确认状态"]
    sheets = [
        ("本月采购明细", display_df(detail, current_cols), "TCurrent"),
        ("历史价格对比", display_df(detail, hist_cols), "THistoryCompare"),
        ("节约金额明细", display_df(detail, saving_cols), "TSavings"),
        ("涨价影响明细", display_df(detail[detail["涨价影响金额"] > 0], increase_cols), "TIncrease"),
        ("品类分析", category, "TCategory"),
        ("供应商分析", supplier, "TSupplier"),
        ("价格异常清单", anomalies, "TAnomaly"),
        ("待人工确认", dq, "TConfirm"),
    ]
    for title, frame, table_name in sheets:
        ws = write_table_sheet(wb, title, frame, table_name)
        if title == "节约金额明细":
            add_savings_formulas(ws)
        add_conditional_formats(ws)
    demo_mode = detail["_源文件"].astype(str).str.contains("演示数据").all()
    method_rows = [
        ["分析期间", month, "本期采购订单仅取订单日期位于分析月份的有效记录"],
        ["统一价格口径", clean_text(rules.get("price_basis", "含税")), "含税价和未税价按税率相互换算；缺少税率时不擅自换算"],
        ["匹配主键", "物料编码+规格+统一单位", "无物料编码仅给候选，不自动匹配"],
        ["上次采购价", "最近一次有效采购单价", "必须早于本次订单日期且税价、单位口径一致"],
        ["12个月加权均价", "Σ(历史单价×历史标准数量)÷Σ历史标准数量", "回溯月数可在分析规则配置修改"],
        ["环比节约金额", "(上次采购价-本次采购价)×本次标准数量", "与其他节约类型分开展示"],
        ["历史基准节约金额", "(12个月加权均价-本次采购价)×本次标准数量", "正式汇总默认采用该口径"],
        ["议价节约金额", "(供应商初始报价-最终成交价)×本次数量", "报价单位默认与订单原单位一致；否则需先补单位字段/换算规则"],
        ["正式汇总节约金额", "仅人工确认=已确认且历史基准节约>0时计入", "演示数据默认全部待确认，因此驾驶舱经确认节约为0"],
        ["涨价影响金额", "MAX((本次采购价-12个月加权均价)×本次数量,0)", "用于识别成本增加，不与节约相抵"],
        ["趋势平均单价限制", "跨物料加权平均单价仅辅助观察", "不同物料和单位混合，不作为正式价格判断"],
        ["数据状态", "演示数据" if demo_mode else "业务输入数据", "演示结果不得当作正式经营结果；业务输入仍须完成人工确认和财务口径复核"],
    ]
    method_df = pd.DataFrame(method_rows, columns=["项目", "口径", "说明"])
    write_table_sheet(wb, "计算口径说明", method_df, "TMethods")
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_simple_report(path: Path, sheet_name: str, df: pd.DataFrame, table_name: str, note: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    ws["A1"] = path.stem
    ws["A1"].font = Font(size=16, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A3"] = note
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100
    write_table_sheet(wb, sheet_name, df, table_name)
    wb.save(path)


def build_pdf(path: Path, month: str, version: int, summary: dict, category: pd.DataFrame, supplier: pd.DataFrame, anomalies: pd.DataFrame, demo_mode: bool):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    styles = getSampleStyleSheet()
    title = ParagraphStyle("cnTitle", parent=styles["Title"], fontName="STSong-Light", fontSize=20, leading=26, alignment=TA_CENTER, textColor=colors.HexColor("#1F4E78"))
    head = ParagraphStyle("cnHead", parent=styles["Heading2"], fontName="STSong-Light", fontSize=13, leading=18, textColor=colors.HexColor("#1F4E78"))
    body = ParagraphStyle("cnBody", parent=styles["BodyText"], fontName="STSong-Light", fontSize=9, leading=14)
    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=15 * mm, leftMargin=15 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
    story = [Paragraph(f"采购月度管理报告{'（演示数据）' if demo_mode else ''}｜{month}", title), Spacer(1, 6 * mm)]
    story.append(Paragraph(f"版本 V{version:03d}｜生成时间 {datetime.now():%Y-%m-%d %H:%M}｜未经人工确认，不得作为正式经营结论。", body))
    kpi_data = [
        ["本月订单数", "采购总金额", "采购物料数", "活跃供应商数", "经确认节约金额", "涨价影响金额", "异常记录数"],
        [summary["本月订单数"], f"{summary['采购总金额']:,.2f}", summary["采购物料数"], summary["活跃供应商数"], f"{summary['经确认节约金额']:,.2f}", f"{summary['涨价影响金额']:,.2f}", summary["异常记录数"]],
    ]
    table = Table(kpi_data, colWidths=[35 * mm] * 7)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9E1F2")), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FB")]),
        ("TOPPADDING", (0, 0), (-1, -1), 7), ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story += [Spacer(1, 5 * mm), table, Spacer(1, 6 * mm), Paragraph("经营结论", head)]
    conclusion = "本报告由演示数据生成，用于验证系统逻辑。" if demo_mode else "本报告由业务输入文件自动生成，关键节约与异常仍须责任人确认。"
    story.append(Paragraph(conclusion + "正式汇总节约金额仅统计人工确认后的近12个月加权平均价节约，三类节约不相加。", body))
    story += [Spacer(1, 4 * mm), Paragraph("品类采购金额", head)]
    cat_rows = [["品类", "采购金额", "金额占比"]] + [[clean_text(r["物料类别"]) or "未分类", f"{r['采购金额']:,.2f}", f"{r['金额占比']:.1%}"] for _, r in category.sort_values("采购金额", ascending=False).iterrows()]
    cat_table = Table(cat_rows, colWidths=[70 * mm, 55 * mm, 45 * mm])
    cat_table.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), "STSong-Light"), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.4, colors.grey), ("ALIGN", (1, 1), (-1, -1), "RIGHT")]))
    story.append(cat_table)
    story += [PageBreak(), Paragraph("重点供应商", head)]
    sup_rows = [["供应商", "采购金额", "订单数", "物料数", "涨价影响"]] + [[clean_text(r["供应商名称"]), f"{r['采购金额']:,.2f}", int(r["订单数"]), int(r["物料数"]), f"{r['涨价影响金额']:,.2f}"] for _, r in supplier.head(10).iterrows()]
    sup_table = Table(sup_rows, colWidths=[80 * mm, 45 * mm, 30 * mm, 30 * mm, 45 * mm])
    sup_table.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), "STSong-Light"), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.4, colors.grey), ("ALIGN", (1, 1), (-1, -1), "RIGHT")]))
    story.append(sup_table)
    story += [Spacer(1, 6 * mm), Paragraph("高风险与重点异常", head)]
    top = anomalies.sort_values(["风险等级", "影响金额"], ascending=[True, False]).head(15) if not anomalies.empty else anomalies
    anom_rows = [["物料", "异常类型", "判断依据", "影响金额", "风险", "建议动作"]]
    for _, r in top.iterrows():
        anom_rows.append([
            Paragraph(clean_text(r["物料名称"]), body), Paragraph(clean_text(r["异常类型"]), body),
            Paragraph(clean_text(r["判断依据"]), body), f"{r['影响金额']:,.2f}",
            clean_text(r["风险等级"]), Paragraph(clean_text(r["建议处理动作"]), body),
        ])
    if len(anom_rows) == 1:
        anom_rows.append(["", "本期无异常", "", "", "", ""])
    anom_table = Table(anom_rows, colWidths=[38 * mm, 38 * mm, 70 * mm, 28 * mm, 20 * mm, 70 * mm], repeatRows=1)
    anom_table.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), "STSong-Light"), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.35, colors.grey), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("ALIGN", (3, 1), (3, -1), "RIGHT")]))
    story.append(anom_table)
    doc.build(story)


def next_version(month_dir: Path, month: str) -> int:
    versions = []
    pattern = re.compile(rf"采购月度分析报告_{re.escape(month)}_V(\d+)\.xlsx$")
    for p in month_dir.glob(f"采购月度分析报告_{month}_V*.xlsx"):
        m = pattern.search(p.name)
        if m:
            versions.append(int(m.group(1)))
    return max(versions, default=0) + 1


def reconcile(detail: pd.DataFrame, category: pd.DataFrame, supplier: pd.DataFrame) -> dict:
    detail_amount = float(detail["含税金额"].fillna(0).sum())
    category_amount = float(category["采购金额"].fillna(0).sum())
    supplier_amount = float(supplier["采购金额"].fillna(0).sum())
    hist_saving = float(detail["历史基准节约金额"].fillna(0).sum())
    saving_positive = float(detail.loc[detail["历史基准节约金额"] > 0, "历史基准节约金额"].sum())
    increase = float(detail["涨价影响金额"].fillna(0).sum())
    checks = {
        "采购总金额_明细": detail_amount,
        "采购总金额_品类": category_amount,
        "采购总金额_供应商": supplier_amount,
        "历史基准节约净额": hist_saving,
        "历史基准正节约分析值": saving_positive,
        "涨价影响金额": increase,
        "品类汇总差异": category_amount - detail_amount,
        "供应商汇总差异": supplier_amount - detail_amount,
    }
    if abs(checks["品类汇总差异"]) > 0.01 or abs(checks["供应商汇总差异"]) > 0.01:
        raise BusinessError(f"汇总金额与明细不一致：{checks}")
    return checks


def verify_workbooks(paths: list[Path], month: str):
    results = []
    for path in paths:
        wb = load_workbook(path, read_only=False, data_only=False)
        formula_count = 0
        formula_errors = []
        chart_count = 0
        for ws in wb.worksheets:
            chart_count += len(ws._charts)
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
                        if any(err in cell.value for err in ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"]):
                            formula_errors.append(f"{ws.title}!{cell.coordinate}:{cell.value}")
        results.append({"文件": str(path), "工作表数": len(wb.sheetnames), "公式数": formula_count, "图表数": chart_count, "公式错误文本数": len(formula_errors)})
        if formula_errors:
            raise BusinessError(f"输出文件存在公式错误文本：{path.name}，示例：{formula_errors[:3]}")
    main = [x for x in results if "采购月度分析报告" in x["文件"]][0]
    if main["公式数"] == 0 or main["图表数"] < 3:
        raise BusinessError("主报告未通过公式/图表完整性检查。")
    return results


def run(month: str):
    started = datetime.now()
    run_id = started.strftime("%Y%m%d_%H%M%S")
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    rules = load_rules()
    if clean_text(rules.get("source_mode", "Excel")).lower() != "excel":
        raise BusinessError("第一阶段仅启用 Excel 数据源。若切换 ERP/Database，请先实现 src/connectors.py。")
    mapping_cfg = load_mapping()
    unit_rules = load_unit_map()
    name_map = load_name_map()

    current_raw, issues1, manifest1 = read_dataset(RAW / "本期采购订单", "采购订单", mapping_cfg, 6)
    history_raw, issues2, manifest2 = read_dataset(RAW / "历史采购订单", "采购订单", mapping_cfg, 6)
    quote_raw, issues3, manifest3 = read_dataset(RAW / "供应商报价", "供应商报价", mapping_cfg, 4)
    receipt_raw, issues4, manifest4 = read_dataset(RAW / "入库记录", "入库记录", mapping_cfg, 3)
    master_raw, issues5, manifest5 = read_dataset(RAW / "物料主数据", "物料主数据", mapping_cfg, 3)
    if current_raw.empty:
        raise BusinessError("本期采购订单目录没有可识别的数据。请放入 Excel，且至少包含订单号、日期、物料、数量、单价、供应商等字段。")

    master = clean_master(master_raw, name_map)
    current_clean, dq1, excluded_current = clean_orders(current_raw, master, unit_rules, name_map, rules, "本期")
    history_clean, dq2, excluded_history = clean_orders(history_raw, master, unit_rules, name_map, rules, "历史")
    quotes = clean_quotes(quote_raw, name_map)
    start = pd.Period(month, "M").start_time
    end = pd.Period(month, "M").end_time
    outside = current_clean[~current_clean["订单日期"].between(start, end, inclusive="both")]
    current = current_clean[current_clean["订单日期"].between(start, end, inclusive="both")].copy()
    if current.empty:
        raise BusinessError(f"本期采购订单中没有订单日期位于 {month} 的有效记录。")
    dq_outside = [{
        "问题类型": "不在分析期间", "文件": r["_源路径"], "工作表": r["_源工作表"], "数据行": r["_源行号"],
        "问题说明": f"订单日期 {r['订单日期']} 不在 {month}，本次未纳入", "建议动作": "确认分析月份或文件放置位置", "人工确认状态": "待确认"
    } for _, r in outside.iterrows()]
    all_orders = pd.concat([history_clean, current_clean], ignore_index=True)
    detail = calculate_baselines(current, all_orders, quotes, rules)

    all_dq = issues1 + issues2 + issues3 + issues4 + issues5 + dq1 + dq2 + dq_outside
    for rec in excluded_current:
        all_dq.append({"问题类型": "订单状态排除", "文件": rec["_源路径"], "工作表": rec["_源工作表"], "数据行": rec["_源行号"], "问题说明": f"订单 {rec['订单号']} 状态“{rec['订单状态']}”已排除", "建议动作": "无需处理；如状态规则变化请修改分析规则配置", "人工确认状态": "已按规则处理"})
    dq_df = pd.DataFrame(all_dq)
    anomalies = detect_anomalies(detail, all_orders, all_dq, rules)
    category, supplier = category_and_supplier(detail)
    trend = monthly_trend(all_orders, month)
    checks = reconcile(detail, category, supplier)

    month_dir = OUT / month
    month_dir.mkdir(parents=True, exist_ok=True)
    version = next_version(month_dir, month)
    suffix = f"{month}_V{version:03d}"
    main_path = month_dir / f"采购月度分析报告_{suffix}.xlsx"
    anomaly_path = month_dir / f"采购价格异常清单_{suffix}.xlsx"
    saving_path = month_dir / f"采购节约金额明细_{suffix}.xlsx"
    dq_path = month_dir / f"数据质量与待确认清单_{suffix}.xlsx"
    pdf_path = month_dir / f"采购月度管理报告_{suffix}.pdf"

    build_main_report(main_path, month, version, detail, all_orders, anomalies, dq_df, category, supplier, trend, rules)
    build_simple_report(anomaly_path, "价格异常清单", anomalies, "TAnomalyOnly", "每条异常均含判断依据、影响金额、风险等级、建议动作和人工确认状态。")
    saving_cols = ["订单号", "订单日期", "物料编码", "标准物料名称", "供应商名称", "标准单位", "标准数量", "标准比较单价", "上次采购价", "近12个月加权平均价", "初始报价", "最终批准报价", "环比节约金额", "历史基准节约金额", "议价节约金额", "节约率", "人工确认状态", "正式汇总节约金额"]
    saving_df = display_df(detail, saving_cols)
    build_simple_report(saving_path, "节约金额明细", saving_df, "TSavingOnly", "三类节约金额分开展示，不直接相加。正式汇总默认采用经人工确认的近12个月加权平均价节约金额。")
    wb_s = load_workbook(saving_path)
    add_savings_formulas(wb_s["节约金额明细"])
    add_conditional_formats(wb_s["节约金额明细"])
    wb_s.save(saving_path)
    build_simple_report(dq_path, "数据质量与待确认", dq_df, "TDQOnly", "无法可靠判断的字段映射、单位换算、缺失字段和候选匹配均在本表列示。")
    summary = {
        "本月订单数": int(detail["订单号"].nunique()),
        "采购总金额": float(detail["含税金额"].fillna(0).sum()),
        "采购物料数": int(detail.loc[detail["物料编码"] != "", "物料编码"].nunique()),
        "活跃供应商数": int(detail["供应商名称"].replace("", pd.NA).nunique()),
        "经确认节约金额": 0.0,
        "涨价影响金额": float(detail["涨价影响金额"].fillna(0).sum()),
        "异常记录数": int(len(anomalies)),
    }
    pdf_status = "未生成"
    demo_mode = detail["_源文件"].astype(str).str.contains("演示数据").all()
    if clean_text(rules.get("generate_pdf", "是")) == "是":
        try:
            build_pdf(pdf_path, month, version, summary, category, supplier, anomalies, demo_mode)
            pdf_status = "已生成"
        except Exception as exc:
            pdf_status = f"生成失败：{exc}"

    output_xlsx = [main_path, anomaly_path, saving_path, dq_path]
    verify = verify_workbooks(output_xlsx, month)
    archive_dir = ARCHIVE / month[:4] / month[5:7] / f"{run_id}_V{version:03d}"
    archive_dir.mkdir(parents=True, exist_ok=True)
    for p in output_xlsx + ([pdf_path] if pdf_path.exists() else []):
        shutil.copy2(p, archive_dir / p.name)

    manifest = manifest1 + manifest2 + manifest3 + manifest4 + manifest5
    log = {
        "运行状态": "成功", "运行开始": started.isoformat(), "运行结束": datetime.now().isoformat(),
        "分析期间": month, "版本": version, "运行ID": run_id, "源文件清单": manifest,
        "记录数量": {"本期原始": len(current_raw), "本期清洗后有效": len(current), "历史清洗后有效": len(history_clean), "报价": len(quotes), "入库": len(receipt_raw), "物料主数据": len(master)},
        "排除数量": {"本期状态排除": len(excluded_current), "历史状态排除": len(excluded_history), "本期非分析期间": len(outside)},
        "异常数量": len(anomalies), "待确认及数据质量数量": len(dq_df),
        "金额一致性检查": checks, "工作簿验证": verify, "PDF状态": pdf_status,
        "输出文件": [str(p) for p in output_xlsx + ([pdf_path] if pdf_path.exists() else [])],
        "归档目录": str(archive_dir),
        "数据声明": "输入文件名含“演示数据”，结果仅用于系统测试，不得作为正式经营分析。" if demo_mode else "业务输入数据已处理；节约、异常和缺失字段仍须人工确认。",
    }
    log_path = LOG_DIR / f"运行日志_{run_id}_{month}_V{version:03d}.json"
    log_path.write_text(json.dumps(log, ensure_ascii=False, indent=2, default=json_value), encoding="utf-8")
    return log


def main():
    parser = argparse.ArgumentParser(description="采购价格与降本月度分析系统")
    parser.add_argument("--month", default=previous_month(), help="分析月份，格式 YYYY-MM")
    args = parser.parse_args()
    month = valid_month(args.month)
    try:
        log = run(month)
        print("运行成功")
        print(f"分析月份：{month}")
        print(f"版本：V{log['版本']:03d}")
        print(f"输出目录：{Path(log['输出文件'][0]).parent}")
        print(f"归档目录：{log['归档目录']}")
        print(f"异常记录：{log['异常数量']}；待确认/数据质量：{log['待确认及数据质量数量']}")
        return 0
    except Exception as exc:
        LOG_DIR.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        error_log = LOG_DIR / f"错误日志_{stamp}_{month}.txt"
        message = f"运行失败\n时间：{datetime.now():%Y-%m-%d %H:%M:%S}\n分析月份：{month}\n错误：{exc}\n\n技术明细：\n{traceback.format_exc()}"
        error_log.write_text(message, encoding="utf-8")
        print(f"[错误] {exc}")
        print(f"详细错误日志：{error_log}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
